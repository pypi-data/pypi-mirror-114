#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2021/7/22 1:24 下午
# @Author  : jianwei.lv

from openpyxl import load_workbook, Workbook
import os
from weimobUI.projectFolder import template_xls_path, template_xls_filename

class extract():
    def __init__(self, xls):
        self.lwb = load_workbook(xls)

    def open_xlsx(self, sheet_name):
        self.ws = self.lwb[sheet_name]
        return self.ws

    """获取每列数据，tuple格式"""
    def get_each_cell(self, start, end):
        self.ws_cell = self.ws[start:end]
        return self.ws_cell

    def close_xlsx(self):
        self.lwb.close()

    def get_cell_str(self, start=15, end=17):
        for each_cell in self.ws_cell:
            for each in each_cell:
                yield str(each)[start:end]

class WR():
    def __init__(self):
        self.wb = Workbook()

    def get_active(self):
        self.ws = self.wb.active

    def create_xlsx_sheet(self, sheet_name):
        self.ws = self.wb.create_sheet(sheet_name)
        return self.ws

    def save_xlsx(self, path):
        self.wb.save(path)

    def wr_xlsx(self, cell, value):
        self.ws[cell].value = value

    def close_xlsx(self):
        self.wb.close()

if __name__ == '__main__':
    a = extract(os.path.join(template_xls_path, template_xls_filename))
    ws = a.open_xlsx(sheet_name='Sheet1')
    ws_cell = a.get_each_cell('A', 'E')
    cell_str = a.get_cell_str()
    a.close_xlsx()

    b = WR()
    b.get_active()

    for each_cell in ws_cell:
        for each in each_cell:
            # print(cell_str.__next__())
            print(each.value)
            b.wr_xlsx(cell_str.__next__(), each.value)

    b.save_xlsx('/Users/lvjianwei/Desktop/workspace/weimob/ui-SDK/sdk/weimobUI/common/test.xlsx')
    b.close_xlsx()
