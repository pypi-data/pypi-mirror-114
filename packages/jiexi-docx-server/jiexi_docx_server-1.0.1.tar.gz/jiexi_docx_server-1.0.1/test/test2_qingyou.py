# -*- coding: utf-8 -*-

from __future__ import unicode_literals

import requests
from openpyxl import load_workbook

excel = load_workbook("E:/test.xlsx")
# 获取sheet：
table = excel.get_sheet_by_name("Sheet1")  # 通过表名获取
# 获取行数和列数：
rows = table.max_row  # 获取行数
cols = table.max_column  # 获取列数
# 获取单元格值：
Data = table.cell(row=row, column=col).value  # 获取表格内容，是从第一行第一列是从1开始的，注意不要丢掉 .value


def save_volume_of_math(source_volume_id_list):
    for source_volume_id in source_volume_id_list:
        res = requests.get(source_volume_id)
        with open(
            r"D:\解析文档\docx_jiexi\数学、英语docx\初中数学箐优网\{}".format(source_volume_id),
            mode="wb",
        ) as f:
            f.write(res.content)
