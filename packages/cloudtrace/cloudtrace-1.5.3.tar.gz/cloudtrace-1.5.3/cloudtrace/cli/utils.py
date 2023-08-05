import os
from abc import ABC, abstractmethod
from multiprocessing.pool import ThreadPool
from zipfile import BadZipFile

import pandas as pd
from openpyxl import load_workbook

def write(rows, outfile, sheet_name=None, **kwargs):
    df = pd.DataFrame(rows)
    for k, v in kwargs.items():
        df[k] = v
    df.sort_values(by='name', inplace=True)
    exists = os.path.exists(outfile)
    if exists:
        try:
            book = load_workbook(outfile)
            try:
                del book[sheet_name]
            except KeyError:
                pass
        except BadZipFile:
            exists = False
    writer = pd.ExcelWriter(outfile, engine='openpyxl')
    if exists:
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    writer.save()

class AbstractCLI(ABC):
    @abstractmethod
    def public_ip_addresses(self, info=None):
        raise NotImplementedError()

    @abstractmethod
    def get_region_names(self):
        raise NotImplementedError()

    def run_all(self, func):
        regnames = self.get_region_names()
        with ThreadPool(min(50, len(regnames))) as pool:
            for _ in pool.imap_unordered(func, regnames):
                pass

    @abstractmethod
    def start_region(self, args):
        raise NotImplementedError()

    @abstractmethod
    def state_region(self, args):
        raise NotImplementedError()

    @abstractmethod
    def stop_region(self, args):
        raise NotImplementedError()

    def write(self, outfile, sheet_name=None, **kwargs):
        rows = []
        self.public_ip_addresses(info=rows)
        df = pd.DataFrame(rows)
        for k, v in kwargs.items():
            df[k] = v
        df.sort_values(by='name', inplace=True)
        exists = os.path.exists(outfile)
        if exists:
            try:
                book = load_workbook(outfile)
                try:
                    del book[sheet_name]
                except KeyError:
                    pass
            except BadZipFile:
                exists = False
        writer = pd.ExcelWriter(outfile, engine='openpyxl')
        if exists:
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.save()